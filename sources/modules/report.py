import os
import tarfile
import re
import xml.etree.ElementTree as etree
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



def unpackXML(dump_file_name):
    tf = tarfile.open( dump_file_name, mode = 'r')
    xml = tf.extractfile('Systemdump.xml')
    tree = etree.parse(xml)
    return tree


def inventoryList(section, worksheet ):
    worksheet.append(["configured", "plugged", "serial number", "hardware revision", "firmware version", "module path"])

    for node in section.iter("Node"):
        row = ["","","","","",""]
        for entry in node.iter():
            if entry.tag == "Module_Status":
                row[0] = entry.attrib.get("Configured")
                row[1] = entry.attrib.get("Plugged")            
            elif entry.tag == 'IO_Information':
                row[2] = entry.attrib.get("Serialnumber")
                row[3] = entry.attrib.get("Hardware_revision")
                row[4] = entry.attrib.get("Firmware_version")
                row[5] = entry.attrib.get("Module_path")

        if re.search("\d[A-Z]", row[0] ): # 'real' B&R hardware ?
            worksheet.append(row)
    return


def report(dump_file_name, type ):
    wb = Workbook() 
    ws = wb.active
    wb.remove(ws)

    tree = unpackXML(dump_file_name)
    root = tree.getroot() 

    for e in root.iter("Section"):
        #print(e.attrib)
        if 'inventory' in type:
            if e.attrib.get("title") == "Hardware":

                ws = wb.create_sheet(title="inventory list")
                inventoryList(e, ws )

    excel_filename = os.path.splitext(dump_file_name)[0] + ".xlsx"
    wb.save(filename = excel_filename)
    return {'result':'Ok'}

